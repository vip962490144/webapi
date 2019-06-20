# -*- coding: utf-8 -*-

from collections import namedtuple

# 将数据从Excel中读取出来，在Python中来处理
from openpyxl import load_workbook  # 可以对已存在的excel进行读写操作
from scripts.handle_config import do_config
from scripts.constants import TEST_DATAS_FILE_PATH


class HandleExcel(object):
    """
    定义处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        # 1、打开excel文件（已存在）
        self.wb = load_workbook(self.filename)

        # 2、定位表单
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active  # 获取第一个表单
        self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        self.Cases = namedtuple("Cases", self.sheet_head_tuple)  # 创建一个元组类
        self.cases_list = []    # 定义一个存放所有Cases命名元组对象

    def get_cases(self):
        """
        获取所有的测试用例
        :return:存放Cases命名元组的列表
        """
        for data in self.ws.iter_rows(min_row=2, values_only=True):  # 每次遍历，返回由某行所有单元格值组成的一个元组
            self.cases_list.append(self.Cases(*data))
        return self.cases_list

    def get_case(self, row):
        """
        获取某一条测试用例
        :param row: 行号
        :return: 一个Cases对象
        """
        if isinstance(row, int) and (1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            print("传入的行号有误，行号应为大于1的整数")

    def write_result(self, row, actual, result):
        """
        将实际值与测试用例执行的结果保存到excel中
        :param row: 保存到哪一行
        :param actual: 实际值
        :param result: 测试用例执行的结果，“Pass”， “Fail”
        :return:
        """
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config("excel", "actual_col"), value=actual)
            other_ws.cell(row=row, column=do_config("excel", "result_col"), value=result)
            other_wb.save(self.filename)
        else:
            print("传入的行号有误，行号应为大于1的整数")


# do_excel = HandleExcel(do_config("file path", "cases_path"))

if __name__ == '__main__':
    # file_name = "cases.xlsx"
    one_excel = HandleExcel(filename=TEST_DATAS_FILE_PATH)
    cases = one_excel.get_cases()
    print(cases)
